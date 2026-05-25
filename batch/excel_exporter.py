from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from batch.batch_runner import (
    SCORE_FRACTION, SCORE_NUMERIC,
    STATUS_ERROR, STATUS_MISSING, STATUS_OK,
    QuestionConfig, QuestionResult, StudentResult,
)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
_C_HDR_BG   = "2F5496"   # dark blue  — header background
_C_HDR_FG   = "FFFFFF"   # white      — header text
_C_PASS     = "C6EFCE"   # light green
_C_PARTIAL  = "FFEB9C"   # light yellow
_C_FAIL     = "FFC7CE"   # light red
_C_MISSING  = "D9D9D9"   # light grey
_C_ERROR    = "F4B8B8"   # salmon
_C_TOTAL_BG = "DCE6F1"   # pale blue  — Total / % columns
_C_ODD_ROW  = "F2F2F2"   # very light grey for alternating rows


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Exporter
# ---------------------------------------------------------------------------

class ExcelExporter:
    """
    Generates a colour-coded .xlsx workbook from a list of StudentResult objects.

    Columns:  Student | Q1 | Q2 | … | Total (X/N) | %
    Rows:     one per student, alternating background for readability
    """

    def export(
        self,
        results: list[StudentResult],
        questions: list[QuestionConfig],
        output_path: Path,
        score_format: str = SCORE_FRACTION,
    ) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marks"
        ws.freeze_panes = "B2"   # freeze header row and student-name column

        n_q = len(questions)
        total_col  = n_q + 2     # 1-indexed: col 1 = student, 2..n_q+1 = questions
        percent_col = total_col + 1

        # ── Header row ───────────────────────────────────────────────────
        headers = (
            ["Student"]
            + [q.label for q in questions]
            + ["Total", "%"]
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill      = _fill(_C_HDR_BG)
            cell.font      = Font(color=_C_HDR_FG, bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = _thin_border()
        ws.row_dimensions[1].height = 22

        # ── Data rows ────────────────────────────────────────────────────
        for row_idx, student in enumerate(results, 2):
            odd = (row_idx % 2 == 0)

            # Student name
            name_cell = ws.cell(row=row_idx, column=1, value=student.label)
            name_cell.alignment = Alignment(horizontal="left", vertical="center",
                                            wrap_text=False)
            name_cell.border = _thin_border()
            if odd:
                name_cell.fill = _fill(_C_ODD_ROW)

            # Question scores
            for col_idx, q in enumerate(questions, 2):
                q_res = student.question_results.get(q.label)
                value = q_res.format(score_format) if q_res else "—"

                # For pure-numeric format store as int so Excel can SUM
                if score_format == SCORE_NUMERIC and q_res and q_res.status == STATUS_OK:
                    value = q_res.passed

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = _thin_border()
                cell.fill      = self._score_fill(q_res)

            # Total column  (always shown as "X / N")
            total_cell = ws.cell(
                row=row_idx, column=total_col,
                value=f"{student.total_passed} / {student.grand_total}",
            )
            total_cell.alignment = Alignment(horizontal="center", vertical="center")
            total_cell.border    = _thin_border()
            total_cell.fill      = _fill(_C_TOTAL_BG)
            total_cell.font      = Font(bold=True)

            # Percentage column
            pct = student.overall_ratio * 100
            pct_cell = ws.cell(
                row=row_idx, column=percent_col,
                value=f"{pct:.1f}%",
            )
            pct_cell.alignment = Alignment(horizontal="center", vertical="center")
            pct_cell.border    = _thin_border()
            pct_cell.fill      = self._ratio_fill(student.overall_ratio)
            pct_cell.font      = Font(bold=True)

        # ── Summary row ──────────────────────────────────────────────────
        if results:
            self._add_summary_row(ws, results, questions, n_q, total_col, percent_col)

        # ── Column widths ────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 36
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        wb.save(output_path)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _add_summary_row(
        self, ws, results, questions, n_q, total_col, percent_col
    ) -> None:
        row = ws.max_row + 2   # leave a blank row

        # Label
        lbl = ws.cell(row=row, column=1, value="Class Average")
        lbl.font   = Font(bold=True, italic=True)
        lbl.border = _thin_border()
        lbl.fill   = _fill(_C_TOTAL_BG)

        # Per-question averages
        for col_idx, q in enumerate(questions, 2):
            ok_results = [
                s.question_results[q.label]
                for s in results
                if q.label in s.question_results
                and s.question_results[q.label].status == STATUS_OK
                and s.question_results[q.label].total > 0
            ]
            if ok_results:
                avg = sum(r.ratio for r in ok_results) / len(ok_results)
                cell = ws.cell(row=row, column=col_idx, value=f"{avg:.0%}")
            else:
                cell = ws.cell(row=row, column=col_idx, value="—")
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(bold=True, italic=True)
            cell.border    = _thin_border()
            cell.fill      = _fill(_C_TOTAL_BG)

        # Overall average %
        ratios = [s.overall_ratio for s in results if s.grand_total > 0]
        overall = sum(ratios) / len(ratios) if ratios else 0
        for col in (total_col, percent_col):
            cell = ws.cell(row=row, column=col,
                           value=f"{overall:.1f}%" if col == percent_col
                           else f"avg {overall:.0%}")
            cell.alignment = Alignment(horizontal="center")
            cell.font      = Font(bold=True, italic=True)
            cell.border    = _thin_border()
            cell.fill      = _fill(_C_TOTAL_BG)

    @staticmethod
    def _score_fill(q_res: QuestionResult | None) -> PatternFill:
        if q_res is None or q_res.status == STATUS_MISSING:
            return _fill(_C_MISSING)
        if q_res.status == STATUS_ERROR:
            return _fill(_C_ERROR)
        return ExcelExporter._ratio_fill(q_res.ratio)

    @staticmethod
    def _ratio_fill(ratio: float) -> PatternFill:
        if ratio >= 0.8:
            return _fill(_C_PASS)
        if ratio >= 0.4:
            return _fill(_C_PARTIAL)
        return _fill(_C_FAIL)
